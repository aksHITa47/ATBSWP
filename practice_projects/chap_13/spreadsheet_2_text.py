import openpyxl, os
from openpyxl import Workbook
from openpyxl.utils.cell import get_column_letter

wb = openpyxl.load_workbook('/home/akki/example.xlsx')
ws = wb['Sheet1']

for col in range(1, ws.max_column+1):
    file = open(f'{col}_column_text.txt','a')
    for row in range(1, ws.max_row+1):
        file.write(str(ws.cell(row=row, column=col).value)+'\n')

    file.close()