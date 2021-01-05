import os
import openpyxl
from openpyxl import Workbook
def blank(start,count,filename):
    wb1 = openpyxl.load_workbook(filename)
    ws1 = wb1.active
    wb2 = openpyxl.Workbook()
    ws2 = wb2['Sheet']
    max_ro = ws1.max_row
    max_col = ws1.max_column
    for i in range(1,start):
        for j in range(1,max_col+1):
            c = ws1.cell(row = i, column = j) 
            ws2.cell(row = i, column = j).value = c.value
    for i in range(start,max_ro+1):
        for j in range(1,max_col+1):
            c = ws1.cell(row = i, column = j) 
            ws2.cell(row = i+count, column = j).value = c.value
    wb2.save(filename)
N=int(input("Enter the row where from blank insertion should start:\n"))
M=int(input("Enter the no of blank rows to be inserted:\n"))
fileName=input("Enter the file:\n")
blank(N,M,fileName)




