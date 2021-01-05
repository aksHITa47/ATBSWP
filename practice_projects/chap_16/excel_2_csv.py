import openpyxl
import csv
import os
from openpyxl import Workbook
from pathlib import Path
#print(os.listdir('.'))
path=Path.cwd()
for foldername, subfolders, filenames in os.walk(path):
    for excelFile in filenames:
        if not excelFile.endswith('.xlsx'):
            continue
        xpath=Path(foldername,excelFile)
        wb=openpyxl.load_workbook(xpath)
        for sheets in wb.get_sheet_names():
            sheetname = wb.get_sheet_by_name(sheets)
            csvfile=open(excelFile[:-5]+'_'+f'{sheets}'+'.csv','w',newline='')
            csvw=csv.writer(csvfile)
            for rowNum in range(1, sheetname.max_row + 1):
                rowData = []
                for colNum in range(1, sheetname.max_column + 1):
                    rowData.append(sheetname.cell(row=rowNum,column=colNum).value)
                for row in rowData:
                    csvw.writerow(row)
            csvfile.close()
        wb.close()
                
              
